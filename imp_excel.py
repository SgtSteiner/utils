# -*- coding: utf8 -*-

import os
from openpyxl import load_workbook
import logging
import tkinter as tk
from tkinter import scrolledtext
from tkinter import ttk
from tkinter import messagebox as mBox
from tkinter import filedialog

MASTER_EXCEL = "master.xlsx"


class Application(ttk.Frame):

    def __init__(self, win):
        super().__init__(win)

        win.geometry("600x420")                # Tamaño de la ventana
        self.place(relwidth=1, relheight=1)         # Ajustamo el frame al tamaño de la ventana

        win.title("Importación de CV ISBAN")

        self.current_directory = os.getcwd()       # Obtiene el directorio actual

        self.labelframe_Top = ttk.LabelFrame(self)
        self.labelframe_Top.place(x=5, y=5, relwidth=0.98, height=95)
        self.labelTop = tk.Label(self.labelframe_Top,
                                 text="Importación de archivos CV ISBAN",
                                 font="bold")
        self.labelTop.pack()
        self.labelPath = tk.Label(self.labelframe_Top,
                                  text=self.current_directory)
        self.labelPath.pack()
        self.plantillaCV = tk.StringVar()
        self.rbSantec = ttk.Radiobutton(self.labelframe_Top, text="España", variable=self.plantillaCV, value="esp")
        self.rbSantec.place(x=500, y=30)
        self.rbIsban = ttk.Radiobutton(self.labelframe_Top, text="ISBAN", variable=self.plantillaCV, value="isban")
        self.rbIsban.place(x=500, y=50)
        self.plantillaCV.set("isban")

        self.ruta_button = ttk.Button(self.labelframe_Top,
                                      text="Seleccionar directorio",
                                      command=self.directory_button_clicked)
        self.ruta_button.pack()

        self.labelframe_arch = ttk.LabelFrame(self, text="Archivos")
        self.labelframe_arch.place(x=5, y=110, relwidth=0.98, height=80)
        self.labelImport = ttk.Label(self.labelframe_arch, text="Procesando: ")
        self.labelImport.place(x=5, y=5)
        self.labelCV = ttk.Label(self.labelframe_arch)
        self.labelCV.place(x=80, y=5)
        self.progressbar = ttk.Progressbar(self.labelframe_arch, length=575, maximum=101)
        self.progressbar.place(x=5, y=30)

        self.labelframe_Detalle = ttk.LabelFrame(self, text="Detalle")
        self.labelframe_Detalle.place(x=5, y=195, relwidth=0.98, height=180)
        self.scr_Detalle = scrolledtext.ScrolledText(self.labelframe_Detalle, width=78, height=10,
                                                     font=('courier', 8, 'normal'))
        self.scr_Detalle.pack()

        self.inicio_button = ttk.Button(self, text="Inicio", command=self.inicio_button_clicked)
        self.inicio_button.place(x=275, y=385)

    def directory_button_clicked(self):
        """ Abre el cuadro de dialogo de seleccionar directorio
        """
        directory = filedialog.askdirectory(initialdir=self.current_directory)
        if directory:
            self.current_directory = directory
            self.labelPath["text"] = self.current_directory
            os.chdir(self.current_directory)
        main_win.update()

    def inicio_button_clicked(self):
        """ Inicia el proceso de importación de archivos
        """

        if self.inicio_button["text"] == "Salir":
            self.quit()
            return

        # Define la configuración del archivo de log
        logging.basicConfig(filename="import.log",
                            level=logging.INFO,
                            format="%(asctime)s:%(levelname)s:%(message)s",
                            filemode="w",
                            )

        logging.info("-- INICIO DEL PROCESO --")
        logging.info("Abriendo archivo excel master {0}".format(MASTER_EXCEL))
        try:
            file_path = os.path.join(self.current_directory, MASTER_EXCEL)
            wb = load_workbook(file_path)
        except FileNotFoundError:
            logging.error("No existe el archivo {0}".format(MASTER_EXCEL), exc_info=True)
            mBox.showerror("Archivo inexistente",
                           "No se encuentra el archivo {0}\nPor favor, selecciona el directorio correcto".format(
                               MASTER_EXCEL))
            return
        logging.info("Obteniendo listado de archivos del directorio")
        archivos_dir = os.listdir(".")
        archivos_excel = []

        # Filtramos aquellos que tienen la extensión .xlsm
        for archivo in archivos_dir:
            if archivo[-5:].upper() == ".XLSM":
                archivos_excel.append(archivo)

        for archivo in archivos_excel:
            self.act_progress(cv_name=archivo, estado_cv=100 / len(archivos_excel))
            logging.info("Abriendo {0}".format(archivo))

            file = load_workbook(archivo)

            # Lectura de datos del candidato
            self.act_progress(cv_name=archivo, cualif_name="Leyendo Datos Generales")
            datos_generales = self.read_seleccion(file, "Datos Generales", 'A4', 'L4')
            self.act_progress(cv_name=archivo, cualif_name="Leyendo Experiencia Laboral")
            datos_experiencia = self.read_seleccion(file, "Experiencia Laboral", 'B7', 'G49')
            self.act_progress(cv_name=archivo, cualif_name="Leyendo Cualificaciones")
            con_funcional = self.read_seleccion(file, "Catálogo Cualificaciones", 'B10', 'E298')
            con_tecnico = self.read_seleccion(file, "Catálogo Cualificaciones", 'H10', 'J108')
            con_prod_santander = self.read_seleccion(file, "Catálogo Cualificaciones", 'M12', 'N52')
            con_prod_mercado = self.read_seleccion(file, "Catálogo Cualificaciones", 'Q12', 'R59')
            if self.plantillaCV == "esp":
                con_perfil = self.read_seleccion(file, "Catálogo Cualificaciones", 'U10', 'W126')
            con_idiomas = self.read_seleccion(file, "Catálogo Cualificaciones", 'Z10', 'AA18')

            # Añade los Datos Generales al archivo master
            self.act_progress(cv_name=archivo, cualif_name="Escribiendo Datos Generales")
            self.write_datos_generales(wb, datos_generales)
            nombre_candidato = datos_generales[0][0] + " " + datos_generales[0][1]
            # Añade los Datos de Experiencia al archivo master
            self.act_progress(cv_name=archivo, cualif_name="Escribiendo Experiencia Laboral")
            self.write_experiencia(wb, datos_experiencia, nombre_candidato)
            # Añade los Datos de Cualificación al archivo master
            self.act_progress(cv_name=archivo, cualif_name="Escribiendo Cualificaciones")
            self.write_cualificacion(wb, con_funcional, "Funcional", nombre_candidato)
            self.write_cualificacion(wb, con_tecnico, "Técnico", nombre_candidato)
            self.write_cualificacion(wb, con_prod_santander, "Prod_Santander", nombre_candidato)
            self.write_cualificacion(wb, con_prod_mercado, "Prod_Mercado", nombre_candidato)
            if self.plantillaCV == "esp":
                self.write_cualificacion(wb, con_perfil, "Perfil", nombre_candidato)
            self.write_cualificacion(wb, con_idiomas, "Idiomas", nombre_candidato)

        logging.info("Cerrando archivo excel master {0}".format(MASTER_EXCEL))
        try:
            wb.save(MASTER_EXCEL)
        except PermissionError:
            logging.error('Fallo al grabar el archivo excel. Archivo ocupado o Permiso denegado', exc_info=True)
            raise
        self.inicio_button["text"] = "Salir"
        logging.info("Archivo cerrado")
        logging.info("-- FIN DEL PROCESO --")
        mBox.showinfo("Proceso finalizado",
                      "Importación realizada con éxito\nImportados {0} archivos".format(len(archivos_excel)))

    def quit(self):
        main_win.quit()
        main_win.destroy()
        exit()

    def act_progress(self, cv_name, estado_cv=None, cualif_name=None):
        """ Actualiza el progreso
            :param cv_name: nombre del cv en procesamiento
            :param estado_cv: cantidad de avance de la barra de progreso
            :param cualif_name: nombre de la cualificación en procesamiento
            """
        self.labelCV["text"] = cv_name
        if estado_cv is not None:
            self.progressbar.step(estado_cv)
        if cualif_name is not None:
            self.scr_Detalle.insert(tk.INSERT, cv_name + " " + cualif_name + "\n")
            self.scr_Detalle.see(tk.END)
        main_win.update()

    def read_seleccion(self, wb, sheet_name, ini, fin):
        """ Devuelve una lista con los valores de las celdas de una fila, según la selección
            :param wb: objeto workbook
            :param sheet_name: nombre de la hoja
            :param ini: celda inicio de la selección
            :param fin: celda fin de la selección
            :return: lista con los valores de las celdas
            """
        logging.info("Obteniendo datos - %s", sheet_name)
        cualif = []
        ws = wb(sheet_name)
        # Importamos el conocimiento funcional
        selection = ws[ini:fin]
        for rows in selection:
            datos = []
            for column in rows:
                datos.append(column.value)
            cualif.append(datos)

        return cualif

    def write_datos_generales(self, wb, datos_gen):
        """ Añade los datos generales del candidato al archivo excel
            :param wb: objeto workbook
            :param datos_gen: lista de datos generales
            """
        logging.info("Escribiendo datos - Datos Generales")
        ws = wb("Datos Generales")
        for dato_gen in datos_gen:
            logging.debug("Escribiendo datos - Datos Generales: {0}".format(dato_gen))
            col = 1
            row = ws.max_row + 1
            for dato in dato_gen:
                ws.cell(row=row, column=col).value = dato
                col += 1

    def write_experiencia(self, wb, datos_exp, candidato):
        """ Añade los datos de experiencia del candidato al archivo excel
            :param wb: objeto workbook
            :param datos_exp: lista de datos de experiencia
            :param candidato: nombre del candidato
            """
        logging.info("Escribiendo datos - Experiencia")
        ws = wb("Experiencia")
        for proyecto in datos_exp:
            if proyecto[0] is not None:
                logging.debug("Escribiendo datos - Experiencia: {0}".format(proyecto))
                col = 1
                row = ws.max_row + 1
                ws.cell(row=row, column=col).value = candidato
                for dato in proyecto:
                    col += 1
                    ws.cell(row=row, column=col).value = dato

    def write_cualificacion(self, wb, catalogos, tipo_catalogo, candidato):
        """ Añade los datos de cualificación del candidato al archivo excel
            :param wb: objeto workbook
            :param catalogos: lista de cualificaciones
            :param tipo_catalogo: tipo de catalogo
            :param candidato: nombre del candidato
            """
        logging.info("Escribiendo datos - {0}".format(tipo_catalogo))
        ws = wb("Cualificación")
        conocimiento = ""
        area = ""
        for catalogo in catalogos:
            logging.debug("Escribiendo datos - {0}: {1}".format(tipo_catalogo, catalogo))
            row = ws.max_row + 1
            if tipo_catalogo == "Funcional":
                if catalogo[0] is not None:
                    conocimiento = catalogo[0]
                if catalogo[1] is not None:
                    area = conocimiento + " / " + catalogo[1]
            elif tipo_catalogo == "Técnico" or tipo_catalogo == "Perfil":
                if catalogo[0] is not None:
                    conocimiento = catalogo[0]
            if catalogo[-1] is not None:
                ws.cell(row=row, column=1).value = candidato
                ws.cell(row=row, column=2).value = tipo_catalogo
                if area:
                    ws.cell(row=row, column=3).value = area
                else:
                    ws.cell(row=row, column=3).value = conocimiento
                ws.cell(row=row, column=4).value = catalogo[-2]
                ws.cell(row=row, column=5).value = catalogo[-1]


if __name__ == "__main__":
    main_win = tk.Tk()
    app = Application(main_win)
    app.mainloop()
